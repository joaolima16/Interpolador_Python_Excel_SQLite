from openpyxl import load_workbook
excel = load_workbook("arquivo.xlsx")
import sqlite3 as sqlite


cont = 1
conn = sqlite.connect("Dbflange.db")
cursor = conn.cursor()
arr_dn = []
arr_dFlange = []
arr_dFuro = []
planilha = excel.active
maximo_linha = planilha.max_row
maximo_colunas = planilha.max_column

for i in range(4, maximo_linha + 1):
    arr_dn.append(planilha.cell(row = i, column=2).value)
    arr_dFlange.append(planilha.cell(row = i, column=6).value)
    arr_dFuro.append(planilha.cell(row = i, column =7).value)

for i in range(len(arr_dn)):
    cursor.execute("""INSERT INTO DIMENSOES(dn,diametro_furacao,diametro_furo,id_norma) 
    VALUES(?,?,?,12)""",(arr_dn[i],arr_dFlange[i],arr_dFuro[i]))
    conn.commit()

print(arr_dFlange)
print(arr_dFuro)
print("Dados inseridos")